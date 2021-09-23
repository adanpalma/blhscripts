from pathlib import Path
from shutil import copyfile

from openpyxl import load_workbook

"""Importante la data se obtiene via SQL de RiesgoCol Tabla= CFC_SQT_RIESGOCO 
COLUMNAS: riesgoco_noprestamo, riesgoco_fechasubida, riesgoco_producto,
riesgoco_mtoprestorg,riesgoco_saldoprest,riesgoco_fechadesem 
riesgoco_fechacancela,riesgoco_estatusmora1,riesgoco_estatusmora2,
riesgoco_estatusmora3,riesgoco_portafolio1,riesgoco_portaname1 
riesgoco_portafolio2,riesgoco_portaname2,riesgoco_tipocartera,
riesgoco_prestamotipo,riesgoco_clasificacion,riesgoco_restructurado 
riesgoco_fechareestru ORDER BY 1,2 La misma se exporta via csv y se graba 
luego como excel que es el archivo que lee este proceso adicional, si cambia 
el nombre del archivo debe cambiar  la variable sourcefilename el excel debe 
tener una pestaña nombrada datatape """
# Make a copy of a original source data tape file because
# for some reason openpyxl some times corrupt file

if __name__ == "__main__":
    path = "/Users/macbook/Documents/LaHipotecaria/"
    sourcefilename = "datatapecolfitch062021.xlsx"
    workfilename = "cop_" + sourcefilename
    sourcefilename = path + sourcefilename
    workfilename = path + workfilename
    sourcefilename = Path(sourcefilename)
    workfilename = Path(workfilename)
    copyfile(sourcefilename, workfilename)

    try:
        wb = load_workbook(workfilename)
        wb.create_sheet("Cosechas", 2)
        wsdatatape = wb["datatape"]
        wscosecha = wb["Cosechas"]

        deemdefault = 0
        desemcosecha = balancedefaulted = 0.00
        prestamoant = prestamo = ""
        maxrows = wsdatatape.max_row + 1

        for row in range(2, maxrows):
            prestamo = wsdatatape["A" + str(row)].value
            if prestamo != prestamoant:

                if prestamoant == "":
                    desemcosecha = wsdatatape["D" + str(row)].value
                    prestamoant = prestamo
                    aniodesem = wsdatatape["F" + str(row)].value
                    tipo = wsdatatape["C" + str(row)].value
                else:
                    wscosecha.append(
                            [prestamoant, aniodesem, desemcosecha,
                             balancedefaulted,
                             tipo
                             ]
                            )
                    desemcosecha = wsdatatape["D" + str(row)].value
                    aniodesem = wsdatatape["F" + str(row)].value
                    tipo = wsdatatape["C" + str(row)].value
                    prestamoant = prestamo
                    deemdefault = 0
                    balancedefaulted = 0.0
            else:
                delincuency = wsdatatape["H" + str(row)].value
                if delincuency == "181-210 DAYS":
                    if deemdefault == 0:
                        balancedefaulted = wsdatatape["E" + str(row)].value
                        deemdefault = 1
    finally:
        wb.save(workfilename)
        wb.close()
